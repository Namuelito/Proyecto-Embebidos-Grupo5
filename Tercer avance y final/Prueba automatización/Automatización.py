import tkinter as tk
from tkinter import ttk, messagebox
import serial
from openpyxl import load_workbook, Workbook
from datetime import datetime
import threading

# Configuración
archivo = "Control_Acceso_2.xlsx"
puerto_arduino = "COM5"
baudrate = 9600

# Cargar o crear el archivo Excel
try:
    wb = load_workbook(archivo)
    hoja_usuarios = wb["Usuarios"]
    hoja_eventos = wb["Eventos"]
except FileNotFoundError:
    wb = Workbook()
    hoja_usuarios = wb.active
    hoja_usuarios.title = "Usuarios"
    hoja_usuarios.append(["UID", "Nombre", "Datos Adicionales"])
    hoja_eventos = wb.create_sheet("Eventos")
    hoja_eventos.append(["UID", "Fecha", "Evento"])
    wb.save(archivo)

# Inicializar comunicación serial con Arduino
try:
    arduino = serial.Serial(puerto_arduino, baudrate, timeout=1)
except Exception as e:
    print(f"Error al conectar con Arduino: {e}")
    arduino = None


# Funciones del Sistema
def registrar_usuario(uid, nombre, datos_adicionales):
    """Registrar un nuevo usuario en la hoja Usuarios."""
    hoja_usuarios.append([uid, nombre, datos_adicionales])
    wb.save(archivo)


def registrar_evento(uid):
    """Registrar un evento en la hoja Eventos."""
    fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    hoja_eventos.append([uid, fecha_actual, "Tarjeta detectada"])
    wb.save(archivo)


def buscar_usuario(uid):
    """Buscar un usuario por UID en la hoja Usuarios."""
    for fila in hoja_usuarios.iter_rows(values_only=True):
        if fila[0] == uid:
            return fila
    return None


def procesar_uid(uid):
    """Procesar UID recibido desde Arduino."""
    usuario = buscar_usuario(uid)
    if usuario:
        registrar_evento(uid)
        actualizar_eventos(f"UID {uid} detectado. Usuario: {usuario[1]}")
    else:
        nuevo_usuario(uid)


def nuevo_usuario(uid):
    """Abrir ventana para registrar un nuevo usuario."""
    def guardar_usuario():
        nombre = entry_nombre.get()
        datos = entry_datos.get()
        if nombre:
            registrar_usuario(uid, nombre, datos)
            actualizar_eventos(f"Nuevo usuario registrado: UID {uid}, Nombre {nombre}")
            ventana.destroy()
        else:
            messagebox.showerror("Error", "El nombre no puede estar vacío.")

    ventana = tk.Toplevel(root)
    ventana.title("Registrar Nuevo Usuario")
    ventana.geometry("300x200")

    tk.Label(ventana, text=f"UID: {uid}").pack(pady=5)
    tk.Label(ventana, text="Nombre:").pack(pady=5)
    entry_nombre = tk.Entry(ventana)
    entry_nombre.pack(pady=5)
    tk.Label(ventana, text="Datos adicionales:").pack(pady=5)
    entry_datos = tk.Entry(ventana)
    entry_datos.pack(pady=5)
    tk.Button(ventana, text="Guardar", command=guardar_usuario).pack(pady=10)


def actualizar_eventos(texto):
    """Actualizar la ventana de eventos en tiempo real."""
    text_eventos.insert(tk.END, texto + "\n")
    text_eventos.see(tk.END)


# GUI Principal
root = tk.Tk()
root.title("Sistema de Control de Acceso")
root.geometry("500x400")

# Etiquetas y Botones
frame_usuarios = tk.LabelFrame(root, text="Usuarios", padx=10, pady=10)
frame_usuarios.pack(fill="x", padx=10, pady=5)
btn_ver_usuarios = tk.Button(frame_usuarios, text="Ver Usuarios", command=lambda: mostrar_usuarios())
btn_ver_usuarios.pack(side="left", padx=5)

frame_eventos = tk.LabelFrame(root, text="Eventos", padx=10, pady=10)
frame_eventos.pack(fill="x", padx=10, pady=5)
text_eventos = tk.Text(frame_eventos, height=10, state="normal")
text_eventos.pack(fill="both", padx=5, pady=5)


def mostrar_usuarios():
    """Mostrar los usuarios registrados en una nueva ventana."""
    ventana = tk.Toplevel(root)
    ventana.title("Usuarios Registrados")
    ventana.geometry("400x300")

    tree = ttk.Treeview(ventana, columns=("UID", "Nombre", "Datos"), show="headings")
    tree.heading("UID", text="UID")
    tree.heading("Nombre", text="Nombre")
    tree.heading("Datos", text="Datos Adicionales")

    for fila in hoja_usuarios.iter_rows(values_only=True):
        tree.insert("", tk.END, values=fila)

    tree.pack(fill="both", expand=True)


# Lectura de Datos en Segundo Plano
def leer_datos():
    while True:
        if arduino and arduino.in_waiting > 0:
            uid = arduino.readline().decode().strip()
            if uid:
                procesar_uid(uid)


# Hilo para leer datos
if arduino:
    hilo = threading.Thread(target=leer_datos, daemon=True)
    hilo.start()

# Ejecutar GUI
root.mainloop()
