import serial
from openpyxl import load_workbook
from datetime import datetime

# Configuración
archivo = 'Control_Acceso.xlsx'  # Archivo Excel
puerto_arduino = "COM6"          # Puerto del Arduino
baudrate = 9600                  # Velocidad de comunicación serial

# Cargar el archivo Excel
try:
    wb = load_workbook(archivo)
    hoja_usuarios = wb["Usuarios"]
    hoja_eventos = wb["Eventos"]
    hoja_ultrasonico = wb["Ultrasonico"]
    print("Archivo Excel cargado correctamente.")
except Exception as e:
    print(f"Error al cargar el archivo Excel: {e}")
    exit()

# Inicializar comunicación serial
try:
    arduino = serial.Serial(puerto_arduino, baudrate, timeout=1)
    print(f"Conexión establecida con Arduino en {puerto_arduino}")
except Exception as e:
    print(f"Error al conectar con Arduino: {e}")
    exit()

# Función para registrar eventos ultrasónicos
def registrar_evento_ultrasonico():
    fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    hoja_ultrasonico.append([fecha_actual])
    wb.save(archivo)
    print("Evento registrado en hoja Ultrasonico")

# Función para validar y registrar acceso
def procesar_acceso(uid, tipo):
    acceso = "Acceso denegado"
    # Validar UID en la hoja `Usuarios`
    for fila in hoja_usuarios.iter_rows(values_only=True):
        if fila[0] == uid and fila[3].lower() == "sí" and fila[4].lower() == tipo.lower():
            acceso = tipo
            break

    # Enviar respuesta al Arduino
    arduino.write(acceso.encode())
    print(f"Respuesta enviada al Arduino: {acceso}")

    # Registrar acceso autorizado
    if acceso != "Acceso denegado":
        fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        hoja_eventos.append([uid, fecha_actual, acceso])
        wb.save(archivo)
        print(f"Acceso {acceso} registrado para UID {uid}")

# Bucle principal
try:
    while True:
        if arduino.in_waiting > 0:  # Si hay datos desde Arduino
            data = arduino.readline().decode().strip()
            print(f"Datos recibidos: {data}")

            if " " in data:  # Verificar formato esperado
                try:
                    uid, tipo = data.split()
                    procesar_acceso(uid, tipo)
                except ValueError:
                    print(f"Error en el formato de los datos: {data}")
            elif data == "ULTRASONICO":
                registrar_evento_ultrasonico()
            else:
                print(f"Datos ignorados: {data}")
except KeyboardInterrupt:
    print("Programa detenido manualmente.")
finally:
    arduino.close()
    print("Conexión serial cerrada.")
